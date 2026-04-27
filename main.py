#!/usr/bin/env python3
from __future__ import annotations

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

import argparse
import asyncio
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from config import DEFAULT_COURSE
from engine import (
    tprint, infer_course, existing_lessons,
    LESSON_TEMPLATE_PATH, UNIT_TEMPLATE_PATH, LOGO_PATH, INPUT_FILE_RE, CONCURRENCY_LIMIT,
    process_file, process_unit,
    run_batch, process_excel_cell, write_report,
    extract_unit_title,
)
from ai import parse_sheet_selection


SEP  = "─" * 56
SEP2 = "═" * 56


def _input(prompt: str) -> str:
    return input(prompt).strip()


def _ask_path(prompt: str, must_exist: bool = True, is_dir: bool = False) -> Path | None:
    while True:
        raw = _input(prompt + " (or ENTER to skip, 'q' to quit): ").strip()
        if raw.lower() == "q":
            raise SystemExit("Bye!")
        if raw == "":
            return None
        p = Path(raw.replace('"', "").replace("'", ""))
        if must_exist and not p.exists():
            print(f"  ✗  Path not found: {p}")
            continue
        if is_dir and must_exist and not p.is_dir():
            print(f"  ✗  Not a directory: {p}")
            continue
        return p.resolve()


def _confirm(prompt: str, default: bool = False) -> bool:
    hint = "[Y/n]" if default else "[y/N]"
    raw = _input(f"{prompt} {hint}: ").lower()
    if raw == "":
        return default
    return raw.startswith("y")


def _print_result_table(reports: list[dict[str, Any]]) -> None:
    print()
    print(f"  {'CODE':<18} {'STATUS'}")
    print(f"  {'─'*18} {'─'*28}")
    for r in reports:
        if r["issues"]:
            status = "⚠  " + ", ".join(r["issues"])
        else:
            status = "✓  ok"
        print(f"  {r['lesson_code']:<18} {status}")


def _detect_schema(ws):
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row:
        return None
    headers = [str(h).strip().lower() if h else "" for h in first_row[0]]
    for i, h in enumerate(headers):
        if "tb script" in h:
            return {"type": "tracking", "markdown_col": i}
    return {"type": "legacy", "unit_col": 0, "lesson_col": 1, "markdown_col": 4}


def _extract_tracking_meta(row, schema, row_idx):
    col = schema["markdown_col"]
    if col >= len(row):
        return None
    text = row[col]
    if not text or not str(text).strip():
        return None
    text_str = str(text).strip()
    m = re.search(r"(?i)^\s*lesson\s*(\d+)[\s:]", text_str)
    if not m:
        m = re.search(r"(?i)lesson\s*(\d+)", text_str)
    lesson = m.group(1) if m else str(row_idx)
    m = re.search(r"(?i)unit\s*(\d+)", text_str)
    unit = m.group(1) if m else "1"
    return {"markdown": text_str, "unit": unit, "lesson": lesson}


def _mode_single_file() -> None:
    print(f"\n{SEP}")
    print("  MODE: Single File")
    print(SEP)

    src = _ask_path("  Lesson markdown file", must_exist=True)
    if src is None:
        print("  Cancelled.")
        return

    default_out = src.parent / "review_output"
    raw_out = _input(f"  Output directory [{default_out}]: ")
    output_dir = Path(raw_out).resolve() if raw_out else default_out

    logo = _ask_path("  Logo image (optional)", must_exist=True)
    logo_path = str(logo) if logo else ""

    save_json = _confirm("  Save intermediate JSON?", default=False)

    print(f"\n{SEP}")
    reports = asyncio.run(run_batch([src], output_dir, logo_path, save_json, infer_course(src)))
    _print_result_table(reports)
    print(f"\n  Output → {output_dir}")
    print(SEP)


def _mode_batch_folder() -> None:
    print(f"\n{SEP}")
    print("  MODE: Batch Folder")
    print(SEP)

    folder = _ask_path("  Folder with lesson markdown files", must_exist=True, is_dir=True)
    if folder is None:
        print("  Cancelled.")
        return

    sources = sorted(p for p in folder.iterdir() if p.is_file() and INPUT_FILE_RE.match(p.name))
    if not sources:
        print(f"  ✗  No matching files (Unit*_Lesson*.md) found in {folder}")
        return

    print(f"\n  Found {len(sources)} file(s):")
    for p in sources:
        print(f"    · {p.name}")

    if not _confirm("\n  Proceed?", default=True):
        print("  Cancelled.")
        return

    default_out = folder / "review_output"
    raw_out = _input(f"\n  Output directory [{default_out}]: ")
    output_dir = Path(raw_out).resolve() if raw_out else default_out

    logo = _ask_path("  Logo image (optional)", must_exist=True)
    logo_path = str(logo) if logo else ""

    save_json = _confirm("  Save intermediate JSON?", default=False)

    print(f"\n{SEP}")
    reports = asyncio.run(run_batch(sources, output_dir, logo_path, save_json, infer_course(folder)))
    if len(reports) > 1:
        write_report(output_dir, reports)
    _print_result_table(reports)
    print(f"\n  Output → {output_dir}")
    print(SEP)


def _mode_excel() -> None:
    """Interactive flow: process markdown cells from an Excel file."""
    UNIT_COL   = 1
    LESSON_COL = 2
    MD_COL     = 5
    SKIP_SHEETS = {"Prompt", "CEFR", "BE Levels", "Marketing"}

    print(f"\n{SEP}")
    print("  MODE: Excel File")
    print(SEP)

    excel_path = _ask_path("  Excel file (.xlsx)", must_exist=True)
    if excel_path is None:
        print("  Cancelled.")
        return

    course_name = infer_course(excel_path)

    try:
        wb = load_workbook(str(excel_path), read_only=False, data_only=True)
    except Exception as exc:
        print(f"  ✗  Failed to open Excel file: {exc}")
        return

    sheet_names = wb.sheetnames
    if not sheet_names:
        print("  ✗  No sheets found in workbook.")
        wb.close()
        return

    LESSON_SHEET_RE = re.compile(r"^L(\d+)$", re.I)
    lesson_sheets = [n for n in sheet_names if LESSON_SHEET_RE.match(n)]
    if not lesson_sheets:
        lesson_sheets = [n for n in sheet_names if n not in SKIP_SHEETS]

    print(f"\n  Found {len(lesson_sheets)} lesson sheet(s):")
    for i, name in enumerate(lesson_sheets, 1):
        print(f"    {i}. {name}")

    raw_sel = _input(f"\n  Select sheets (e.g. 1,4 or 1-4 or 'all') [all]: ")
    selected_indices = parse_sheet_selection(raw_sel, len(lesson_sheets))
    if not selected_indices:
        print("  ✗  No valid sheets selected.")
        wb.close()
        return

    selected_names = [lesson_sheets[i] for i in selected_indices]
    print(f"\n  Selected {len(selected_names)} sheet(s):")
    for name in selected_names:
        print(f"    · {name}")

    default_out = excel_path.parent / "output"
    raw_out = _input(f"\n  Output directory [{default_out}]: ")
    output_dir = Path(raw_out).resolve() if raw_out else default_out

    logo = _ask_path("  Logo image (optional)", must_exist=True)
    logo_path = str(logo) if logo else ""

    save_json = _confirm("  Save intermediate JSON?", default=False)

    if not _confirm("\n  Proceed?", default=True):
        print("  Cancelled.")
        wb.close()
        return

    template = LESSON_TEMPLATE_PATH.read_text(encoding="utf-8")
    output_dir.mkdir(parents=True, exist_ok=True)

    skip_existing = os.environ.get("REPORT_SKIP_EXISTING", "1").lower() not in ("0", "false", "no", "")
    existing_codes = existing_lessons(output_dir) if skip_existing else set()
    if skip_existing and existing_codes:
        print(f"  {len(existing_codes)} existing PDF(s) will be skipped.")

    tasks: list[dict[str, Any]] = []
    for sheet_name in selected_names:
        ws = wb[sheet_name]
        schema = _detect_schema(ws)
        if not schema:
            continue

        if schema["type"] == "tracking":
            m = re.search(r"(?i)level\s*(\d+)", str(excel_path.name))
            level = m.group(1) if m else "?"
        else:
            m = LESSON_SHEET_RE.match(sheet_name)
            level = m.group(1) if m else sheet_name

        print(f"  ── Sheet: {sheet_name}  (Level {level}) reading rows...")
        row_count = 0
        skipped_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if schema["type"] == "tracking":
                meta = _extract_tracking_meta(row, schema, row_idx)
                if not meta:
                    continue
                markdown_text = meta["markdown"]
                unit = meta["unit"]
                lesson = meta["lesson"]
            else:
                try:
                    markdown_text = row[schema["markdown_col"]]
                except IndexError:
                    continue
                if not markdown_text or not str(markdown_text).strip():
                    continue
                markdown_text = str(markdown_text).strip()

                try:
                    raw_unit = row[schema["unit_col"]]
                    unit = str(int(raw_unit)) if raw_unit is not None else "?"
                except (TypeError, ValueError):
                    unit = str(raw_unit).strip() if raw_unit else "?"

                try:
                    raw_lesson = row[schema["lesson_col"]]
                    lesson = str(int(raw_lesson)) if raw_lesson is not None else "?"
                except (TypeError, ValueError):
                    lesson = str(raw_lesson).strip() if raw_lesson else "?"

                if not unit or unit.lower() == "none":
                    unit = "?"
                if not lesson or lesson.lower() == "none":
                    lesson = "?"

            lesson_code = f"L{level}U{unit}L{lesson}"
            if skip_existing and lesson_code in existing_codes:
                skipped_count += 1
                continue

            row_count += 1
            tasks.append(dict(
                markdown_text=markdown_text,
                level=level, unit=unit, lesson=lesson,
                output_dir=output_dir, template=template,
                logo_path=logo_path, save_json=save_json,
                course_name=course_name,
            ))

        print(f"    → {row_count} rows queued" + (f", {skipped_count} skipped" if skipped_count else ""))
        if row_count == 0 and skipped_count == 0:
            print(f"    (no valid rows found)")

    wb.close()

    if not tasks:
        print("  No tasks to process.")
        return

    total = len(tasks)
    print(f"\n  Dispatching {total} lessons with asyncio (concurrency {CONCURRENCY_LIMIT})...")

    async def _run_excel_batch() -> list[dict[str, Any]]:
        sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
        done = 0

        async def _wrap(task_dict: dict[str, Any]) -> dict[str, Any]:
            nonlocal done
            async with sem:
                try:
                    result = await process_excel_cell(**task_dict)
                except Exception as exc:
                    code = f"L{task_dict['level']}U{task_dict['unit']}L{task_dict['lesson']}"
                    await tprint(f"  [ERROR] {code}: {exc}")
                    result = {"lesson_code": code, "html": "", "pdf": "", "png": "", "json": "", "issues": [f"error:{exc}"]}
                done += 1
                await tprint(f"  Progress: {done}/{total}")
                return result

        return await asyncio.gather(*[_wrap(t) for t in tasks])

    reports = asyncio.run(_run_excel_batch())

    if len(reports) > 1:
        write_report(output_dir, reports)
    _print_result_table(reports)
    print(f"\n  Output → {output_dir}")
    print(SEP)


def _mode_unit_summary_folder() -> None:
    print(f"\n{SEP}")
    print("  MODE: Unit Summary (Folder)")
    print(SEP)

    folder = _ask_path("  Folder with lesson markdown files", must_exist=True, is_dir=True)
    if folder is None:
        print("  Cancelled.")
        return

    sources = sorted(p for p in folder.iterdir() if p.is_file() and INPUT_FILE_RE.match(p.name))
    if not sources:
        print(f"  ✗  No matching files (Unit*_Lesson*.md) found in {folder}")
        return

    unit_groups: dict[str, list[Path]] = defaultdict(list)
    for p in sources:
        m = INPUT_FILE_RE.match(p.name)
        if m:
            unit_groups[m.group(1)].append(p)

    if not unit_groups:
        print("  ✗  No units found.")
        return

    print(f"\n  Found {len(sources)} file(s) across {len(unit_groups)} unit(s):")
    for u, ps in sorted(unit_groups.items()):
        print(f"    · Unit {u}: {len(ps)} lesson(s)")

    if not _confirm("\n  Proceed?", default=True):
        print("  Cancelled.")
        return

    level_guess = "?"
    for part in reversed(folder.parts):
        m = re.match(r"^L(\d+)$", part, re.I)
        if m:
            level_guess = m.group(1)
            break

    raw_level = _input(f"  Level [{level_guess}]: ")
    level = raw_level if raw_level else level_guess

    default_out = folder / "review_output"
    raw_out = _input(f"\n  Output directory [{default_out}]: ")
    output_dir = Path(raw_out).resolve() if raw_out else default_out

    logo = _ask_path("  Logo image (optional)", must_exist=True)
    logo_path = str(logo) if logo else ""

    print(f"\n{SEP}")

    template = UNIT_TEMPLATE_PATH.read_text(encoding="utf-8")
    output_dir.mkdir(parents=True, exist_ok=True)

    unit_tasks: list[tuple[str, str, str, str, str, str]] = []
    for unit, paths in sorted(unit_groups.items()):
        parts = []
        first_markdown = ""
        for p in sorted(paths):
            text = p.read_text(encoding="utf-8")
            if not first_markdown:
                first_markdown = text
            parts.append(f"<!-- {p.name} -->\n\n{text}")
        combined = "\n\n---\n\n".join(parts)
        unit_code = f"L{level}U{unit}"
        extracted_title = extract_unit_title(first_markdown)
        # Find next unit's first lesson for preview
        next_unit_num = int(unit) + 1
        next_preview_path = folder / f"Unit{next_unit_num}_Lesson1.md"
        next_preview = next_preview_path.read_text(encoding="utf-8") if next_preview_path.exists() else ""
        unit_tasks.append((unit_code, level, unit, combined, extracted_title, next_preview))

    total = len(unit_tasks)
    print(f"  Processing {total} unit(s) with asyncio (concurrency limit {CONCURRENCY_LIMIT})...")

    async def _run() -> list[dict[str, Any]]:
        sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
        done = 0

        async def _wrap(task: tuple[str, str, str, str, str, str]) -> dict[str, Any]:
            nonlocal done
            unit_code, lvl, unit, combined, extracted_title, next_preview = task
            async with sem:
                result = await process_unit(unit_code, combined, output_dir, template, logo_path, lvl, unit, extracted_title, course_name, next_preview)
                done += 1
                await tprint(f"  [{done}/{total}] {unit_code} completed")
                return result

        return await asyncio.gather(*[_wrap(t) for t in unit_tasks])

    reports = asyncio.run(_run())
    if len(reports) > 1:
        write_report(output_dir, reports)
    _print_result_table(reports)
    print(f"\n  Output → {output_dir}")
    print(SEP)


def _mode_unit_summary_excel() -> None:
    """Interactive flow: process unit summaries from an Excel file."""
    UNIT_COL   = 1
    LESSON_COL = 2
    MD_COL     = 5
    SKIP_SHEETS = {"Prompt", "CEFR", "BE Levels", "Marketing"}
    LESSON_SHEET_RE = re.compile(r"^L(\d+)$", re.I)

    print(f"\n{SEP}")
    print("  MODE: Unit Summary (Excel)")
    print(SEP)

    excel_path = _ask_path("  Excel file (.xlsx)", must_exist=True)
    if excel_path is None:
        print("  Cancelled.")
        return

    course_name = infer_course(excel_path)

    try:
        wb = load_workbook(str(excel_path), read_only=False, data_only=True)
    except Exception as exc:
        print(f"  ✗  Failed to open Excel file: {exc}")
        return

    sheet_names = wb.sheetnames
    if not sheet_names:
        print("  ✗  No sheets found in workbook.")
        wb.close()
        return

    lesson_sheets = [n for n in sheet_names if LESSON_SHEET_RE.match(n)]
    if not lesson_sheets:
        lesson_sheets = [n for n in sheet_names if n not in SKIP_SHEETS]

    print(f"\n  Found {len(lesson_sheets)} lesson sheet(s):")
    for i, name in enumerate(lesson_sheets, 1):
        print(f"    {i}. {name}")

    raw_sel = _input(f"\n  Select sheets (e.g. 1,4 or 1-4 or 'all') [all]: ")
    selected_indices = parse_sheet_selection(raw_sel, len(lesson_sheets))
    if not selected_indices:
        print("  ✗  No valid sheets selected.")
        wb.close()
        return

    selected_names = [lesson_sheets[i] for i in selected_indices]
    print(f"\n  Selected {len(selected_names)} sheet(s):")
    for name in selected_names:
        print(f"    · {name}")

    default_out = excel_path.parent / "output"
    raw_out = _input(f"\n  Output directory [{default_out}]: ")
    output_dir = Path(raw_out).resolve() if raw_out else default_out

    logo = _ask_path("  Logo image (optional)", must_exist=True)
    logo_path = str(logo) if logo else ""

    if not _confirm("\n  Proceed?", default=True):
        print("  Cancelled.")
        wb.close()
        return

    template = UNIT_TEMPLATE_PATH.read_text(encoding="utf-8")
    output_dir.mkdir(parents=True, exist_ok=True)

    unit_groups: dict[tuple[str, str], list[str]] = defaultdict(list)
    unit_first: dict[tuple[str, str], str] = {}

    for sheet_name in selected_names:
        ws = wb[sheet_name]
        m = LESSON_SHEET_RE.match(sheet_name)
        level = m.group(1) if m else sheet_name

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                markdown_text = row[MD_COL - 1]
            except IndexError:
                continue
            if not markdown_text or not str(markdown_text).strip():
                continue
            markdown_text = str(markdown_text).strip()

            try:
                raw_unit = row[UNIT_COL - 1]
                unit = str(int(raw_unit)) if raw_unit is not None else "?"
            except (TypeError, ValueError):
                unit = str(raw_unit).strip() if raw_unit else "?"

            if not unit or unit.lower() == "none":
                continue

            key = (level, unit)
            unit_groups[key].append(markdown_text)
            if key not in unit_first:
                unit_first[key] = markdown_text

    wb.close()

    if not unit_groups:
        print("  No valid unit data found.")
        return

    print(f"\n  Found {len(unit_groups)} unit(s) to summarize.")

    unit_tasks: list[tuple[str, str, str, str, str, str]] = []
    for (level, unit), mds in sorted(unit_groups.items()):
        combined = "\n\n---\n\n".join(mds)
        unit_code = f"L{level}U{unit}"
        extracted_title = extract_unit_title(unit_first.get((level, unit), ""))
        # Find next unit's first lesson for preview
        next_unit = str(int(unit) + 1)
        next_preview = unit_first.get((level, next_unit), "")
        unit_tasks.append((unit_code, level, unit, combined, extracted_title, next_preview))

    total = len(unit_tasks)
    print(f"  Processing {total} unit(s) with asyncio (concurrency limit {CONCURRENCY_LIMIT})...")

    async def _run() -> list[dict[str, Any]]:
        sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
        done = 0

        async def _wrap(task: tuple[str, str, str, str, str, str]) -> dict[str, Any]:
            nonlocal done
            unit_code, lvl, unit, combined, extracted_title, next_preview = task
            async with sem:
                result = await process_unit(unit_code, combined, output_dir, template, logo_path, lvl, unit, extracted_title, course_name, next_preview)
                done += 1
                await tprint(f"  [{done}/{total}] {unit_code} completed")
                return result

        return await asyncio.gather(*[_wrap(t) for t in unit_tasks])

    reports = asyncio.run(_run())
    if len(reports) > 1:
        write_report(output_dir, reports)
    _print_result_table(reports)
    print(f"\n  Output → {output_dir}")
    print(SEP)


def _mode_unit_summary() -> None:
    print(f"\n{SEP}")
    print("  MODE: Unit Summary")
    print(SEP)
    print("  Source:")
    print("    1. Folder with markdown files")
    print("    2. Excel file")
    print(SEP)

    choice = _input("  Select source > ")
    if choice == "1":
        _mode_unit_summary_folder()
    elif choice == "2":
        _mode_unit_summary_excel()
    else:
        print("  Cancelled.")


def menu() -> None:
    while True:
        print(f"\n{SEP2}")
        print("  A1 Review Card Generator")
        print(SEP2)
        print("  1. Process a single lesson file")
        print("  2. Process a folder of lesson files (batch)")
        print("  3. Process an Excel file")
        print("  4. Generate unit summary")
        print("  0. Quit")
        print(SEP2)

        choice = _input("  Select > ")

        if choice == "1":
            _mode_single_file()
        elif choice == "2":
            _mode_batch_folder()
        elif choice == "3":
            _mode_excel()
        elif choice == "4":
            _mode_unit_summary()
        elif choice in ("0", "q", "quit", "exit"):
            print("\n  Bye!\n")
            break
        else:
            print("  Invalid choice, please enter 1, 2, 3, 4, or 0.")


def cli_excel(args: Any) -> None:
    """Non-interactive Excel batch processing from CLI arguments."""
    UNIT_COL   = 1
    LESSON_COL = 2
    MD_COL     = 5
    SKIP_SHEETS = {"Prompt", "CEFR", "BE Levels", "Marketing"}
    LESSON_SHEET_RE = re.compile(r"^L(\d+)$", re.I)

    excel_path = Path(args.excel)
    output_dir = Path(args.output_dir).resolve() if args.output_dir else excel_path.parent / "output"
    logo_path  = args.logo or ""
    save_json  = bool(args.save_json)
    level_filter = {str(l).strip() for l in args.levels} if args.levels else set()
    course_name = infer_course(excel_path)

    print(f"  Excel   : {excel_path}")
    print(f"  Output  : {output_dir}")
    print(f"  Logo    : {logo_path or '(none)'}")
    print(f"  Save JSON: {save_json}")
    print(f"  Levels  : {', '.join(sorted(level_filter)) if level_filter else 'all'}")

    wb = load_workbook(str(excel_path), read_only=False, data_only=True)
    sheet_names = wb.sheetnames
    lesson_sheets = [n for n in sheet_names if LESSON_SHEET_RE.match(n)]
    if not lesson_sheets:
        lesson_sheets = [n for n in sheet_names if n not in SKIP_SHEETS]

    selected_names = lesson_sheets

    if not selected_names:
        print("  ✗  No matching sheets to process.")
        wb.close()
        return

    print(f"  Sheets  : {', '.join(selected_names)}")

    template = LESSON_TEMPLATE_PATH.read_text(encoding="utf-8")
    output_dir.mkdir(parents=True, exist_ok=True)

    skip_existing = os.environ.get("REPORT_SKIP_EXISTING", "1").lower() not in ("0", "false", "no", "")
    existing_codes = existing_lessons(output_dir) if skip_existing else set()
    if skip_existing and existing_codes:
        print(f"  {len(existing_codes)} existing PDF(s) will be skipped.")

    tasks: list[dict[str, Any]] = []
    for sheet_name in selected_names:
        ws = wb[sheet_name]
        schema = _detect_schema(ws)
        if not schema:
            continue

        if schema["type"] == "tracking":
            m = re.search(r"(?i)level\s*(\d+)", str(excel_path.name))
            level = m.group(1) if m else "?"
        else:
            m = LESSON_SHEET_RE.match(sheet_name)
            level = m.group(1) if m else sheet_name

        if level_filter and level not in level_filter:
            continue

        row_count = 0
        skipped_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if schema["type"] == "tracking":
                meta = _extract_tracking_meta(row, schema, row_idx)
                if not meta:
                    continue
                markdown_text = meta["markdown"]
                unit = meta["unit"]
                lesson = meta["lesson"]
            else:
                try:
                    markdown_text = row[schema["markdown_col"]]
                except IndexError:
                    continue
                if not markdown_text or not str(markdown_text).strip():
                    continue
                markdown_text = str(markdown_text).strip()

                try:
                    raw_unit = row[schema["unit_col"]]
                    unit = str(int(raw_unit)) if raw_unit is not None else "?"
                except (TypeError, ValueError):
                    unit = str(raw_unit).strip() if raw_unit else "?"

                try:
                    raw_lesson = row[schema["lesson_col"]]
                    lesson = str(int(raw_lesson)) if raw_lesson is not None else "?"
                except (TypeError, ValueError):
                    lesson = str(raw_lesson).strip() if raw_lesson else "?"

                if not unit or unit.lower() == "none":
                    unit = "?"
                if not lesson or lesson.lower() == "none":
                    lesson = "?"

            lesson_code = f"L{level}U{unit}L{lesson}"
            if skip_existing and lesson_code in existing_codes:
                skipped_count += 1
                continue

            row_count += 1
            tasks.append(dict(
                markdown_text=markdown_text,
                level=level, unit=unit, lesson=lesson,
                output_dir=output_dir, template=template,
                logo_path=logo_path, save_json=save_json,
                course_name=course_name,
            ))

        print(f"    {sheet_name}: {row_count} rows queued" + (f", {skipped_count} skipped" if skipped_count else ""))

    wb.close()

    if not tasks:
        print("  No tasks to process.")
        return

    total = len(tasks)
    print(f"\n  Dispatching {total} lessons with asyncio (concurrency {CONCURRENCY_LIMIT})...")

    async def _run_excel_batch() -> list[dict[str, Any]]:
        sem = asyncio.Semaphore(CONCURRENCY_LIMIT)
        done = 0

        async def _wrap(task_dict: dict[str, Any]) -> dict[str, Any]:
            nonlocal done
            async with sem:
                try:
                    result = await process_excel_cell(**task_dict)
                except Exception as exc:
                    code = f"L{task_dict['level']}U{task_dict['unit']}L{task_dict['lesson']}"
                    await tprint(f"  [ERROR] {code}: {exc}")
                    result = {"lesson_code": code, "html": "", "pdf": "", "png": "", "json": "", "issues": [f"error:{exc}"]}
                done += 1
                await tprint(f"  Progress: {done}/{total}")
                return result

        return await asyncio.gather(*[_wrap(t) for t in tasks])

    reports = asyncio.run(_run_excel_batch())

    if len(reports) > 1:
        write_report(output_dir, reports)
    _print_result_table(reports)
    print(f"\n  Output → {output_dir}")
    print(SEP)


def main() -> None:
    parser = argparse.ArgumentParser(description="A1 Review Card Generator")
    parser.add_argument("--excel", type=Path, help="Path to Excel file")
    parser.add_argument("--output-dir", type=Path, help="Output directory")
    parser.add_argument("--logo", type=str, default="", help="Path to logo image")
    parser.add_argument("--save-json", action="store_true", help="Save intermediate JSON")
    parser.add_argument("--levels", nargs="+", help="Levels to process")
    args = parser.parse_args()

    if args.excel:
        cli_excel(args)
    else:
        menu()


if __name__ == "__main__":
    main()
