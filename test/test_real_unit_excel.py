import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from engine import process_unit, extract_unit_title
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
EXCEL_PATH = ROOT / "KSA BE 20260324.xlsx"
TEMPLATE_PATH = ROOT / "template" / "unit_template.html"
LOGO_PATH = ROOT / "51talklogo.png"
OUTPUT_DIR = ROOT / "test_output_real_excel"

UNIT_COL = 1
MD_COL = 5


def read_unit_markdowns(sheet_name: str, target_unit: str):
    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[sheet_name]
    mds = []
    next_preview = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            u = str(int(row[UNIT_COL - 1])) if row[UNIT_COL - 1] is not None else ""
        except (TypeError, ValueError):
            u = str(row[UNIT_COL - 1]).strip() if row[UNIT_COL - 1] else ""
        md = row[MD_COL - 1]
        if not u or not md or not str(md).strip():
            continue
        if u == target_unit:
            mds.append(str(md).strip())
        if u == str(int(target_unit) + 1) and not next_preview:
            next_preview = str(md).strip()
    wb.close()
    return mds, next_preview


async def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    logo = str(LOGO_PATH) if LOGO_PATH.exists() else ""

    mds, next_preview = read_unit_markdowns("L1", "1")
    print(f"Unit 1 lessons: {len(mds)}, next preview available: {bool(next_preview)}")
    if not mds:
        print("No data found")
        return

    combined = "\n\n---\n\n".join(mds)
    unit_title = _extract_unit_title(mds[0])

    result = await process_unit(
        unit_code="L1U1",
        combined_markdown=combined,
        output_dir=OUTPUT_DIR,
        template=template,
        logo_path=logo,
        level="1",
        unit="1",
        unit_title=unit_title,
        course_name="Business English",
        next_unit_preview_md=next_preview,
    )
    print("Result:", result)


if __name__ == "__main__":
    asyncio.run(main())
